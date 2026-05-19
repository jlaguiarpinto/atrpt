# presentation/correio/correio_controller.py

import re
import unicodedata
import logging
from collections import Counter
from email.utils import getaddresses, parseaddr, parsedate_to_datetime
from pathlib import Path
from infrastructure.email.imap_client import ImapClient
from infrastructure.persistence.aprovisionamento.email_fornecedor_map_repository import (
    EmailFornecedorMapRepository,
)

logger = logging.getLogger(__name__)


_DOUBLE_TLDS = frozenset(["co.uk", "com.br", "com.pt", "co.pt", "org.pt", "net.pt"])


def _base_domain(addr: str) -> str:
    """
    Extrai o nome de base do domínio ignorando subdomínios e TLD.
    'noreply@mail.kone.pt' → 'kone'   |   'faturas@edp.pt' → 'edp'
    """
    at = addr.rfind("@")
    if at < 0:
        return ""
    parts = addr[at + 1:].lower().split(".")
    if len(parts) >= 3 and ".".join(parts[-2:]) in _DOUBLE_TLDS:
        return parts[-3]
    return parts[-2] if len(parts) >= 2 else ""


def _prefixo_fornecedor(nome: str) -> str:
    """Primeiros 5 caracteres alfanuméricos uppercase do nome (sem acentos)."""
    norm = unicodedata.normalize("NFD", nome)
    limpo = "".join(c for c in norm if unicodedata.category(c) != "Mn")
    alnum = re.sub(r"[^A-Z0-9]", "", limpo.upper())
    return (alnum + "XXXXX")[:5]


def _tipo_doc(n_fatura: str | None) -> str:
    """'nc' para notas de crédito/débito; 'fat' para tudo o resto."""
    if n_fatura:
        m = re.match(r"^(NC|ND)\b", n_fatura.strip(), re.IGNORECASE)
        if m:
            return "nc"
    return "fat"

_STOP = {
    "de", "a", "o", "e", "do", "da", "dos", "das", "em", "um", "uma",
    "para", "com", "por", "que", "se", "no", "na", "nos", "nas", "ao",
    "as", "os", "não", "mais", "como", "mas", "ou", "já", "foi",
    "ser", "ter", "este", "esta", "isso", "essa", "pelo", "pela",
    "seu", "sua", "seus", "suas", "num", "numa", "nem", "também",
    "re", "fw", "fwd", "enc", "rv", "aw", "res", "enc",
}

def _normalizar_enderecos(raw: str) -> list[str]:
    pairs = getaddresses([raw])
    return [addr.strip().lower() for _, addr in pairs if addr.strip()]

def _palavras_assunto(assunto: str) -> list[str]:
    assunto = re.sub(r'^\s*(re|fw|fwd|enc|rv|aw|res)\s*:\s*', '', assunto, flags=re.IGNORECASE)
    tokens = re.findall(r'\b\w{3,}\b', assunto.lower(), re.UNICODE)
    return [t for t in tokens if t not in _STOP and not t.isdigit()]


class CorreioController:

    def __init__(self, gui, imap_client: ImapClient,
                 fornecedor_repo=None,
                 faturacao_repo=None,
                 pasta_faturas: Path | None = None,
                 nif_proprio: str | None = None):
        self.gui              = gui
        self._imap            = imap_client
        self._fornecedor_repo = fornecedor_repo
        self._faturacao_repo  = faturacao_repo
        self.pasta_faturas    = pasta_faturas
        self._nif_proprio     = nif_proprio
        self._pasta_actual: str | None = None
        self._email_map_repo  = (
            EmailFornecedorMapRepository(fornecedor_repo.db_path)
            if fornecedor_repo else None
        )

    # ------------------------------------------------------------------

    def conectar(self) -> bool:
        try:
            self._imap.connect()
            return True
        except Exception as e:
            logger.exception("Falha ao ligar IMAP")
            if self.gui:
                self.gui.informuser("Erro de ligação", str(e), tipo="error")
            return False

    def conectar_se_necessario(self) -> bool:
        if self._imap.is_connected:
            return True
        return self.conectar()

    def desligar(self):
        self._imap.disconnect()

    # ------------------------------------------------------------------

    def listar_pastas(self) -> list[str]:
        try:
            return self._imap.listar_pastas()
        except Exception:
            logger.exception("Erro ao listar pastas")
            return []

    def listar_pastas_com_contagem(self) -> list[tuple[str, int]]:
        try:
            return self._imap.status_pastas()
        except Exception:
            logger.exception("Erro ao obter contagem de pastas")
            return []

    def _stats_pasta(self, pasta: str, limite: int) -> dict:
        """Counters brutos — uso interno."""
        try:
            cabecalhos = self._imap.listar_cabecalhos_analise(pasta, limite)
        except Exception:
            logger.exception("Erro ao analisar pasta '%s'", pasta)
            return {"total": 0, "num_anexos": 0,
                    "emissores": Counter(), "destinatarios": Counter(), "palavras": Counter()}

        emissores     = Counter()
        destinatarios = Counter()
        palavras      = Counter()
        num_anexos    = 0

        for h in cabecalhos:
            for addr in _normalizar_enderecos(h["de"]):
                emissores[addr] += 1
            for addr in _normalizar_enderecos(h["para"]):
                destinatarios[addr] += 1
            for w in _palavras_assunto(h["assunto"]):
                palavras[w] += 1
            if h.get("tem_anexo"):
                num_anexos += 1

        return {
            "total":         len(cabecalhos),
            "num_anexos":    num_anexos,
            "emissores":     emissores,
            "destinatarios": destinatarios,
            "palavras":      palavras,
        }

    def analisar_pasta(self, pasta: str, limite: int = 500) -> dict:
        raw = self._stats_pasta(pasta, limite)
        return {
            "total":          raw["total"],
            "num_anexos":     raw["num_anexos"],
            "emissores":      raw["emissores"].most_common(25),
            "destinatarios":  raw["destinatarios"].most_common(25),
            "palavras_chave": raw["palavras"].most_common(5),
        }

    def analisar_todas_pastas(self, limite: int = 500, callback=None) -> dict:
        pastas = sorted(self._imap.status_pastas(), key=lambda x: x[0].lower())
        por_pasta       = []
        palavras_global = Counter()

        for i, (pasta, _) in enumerate(pastas, 1):
            if callback:
                callback(i, len(pastas), pasta)
            raw = self._stats_pasta(pasta, limite)
            palavras_global.update(raw["palavras"])
            por_pasta.append({
                "pasta":         pasta,
                "total":         raw["total"],
                "num_anexos":    raw["num_anexos"],
                "emissores":     raw["emissores"].most_common(25),
                "destinatarios": raw["destinatarios"].most_common(25),
                "palavras_chave": raw["palavras"].most_common(5),
            })

        return {
            "por_pasta":       por_pasta,
            "palavras_global": palavras_global.most_common(50),
        }

    # ------------------------------------------------------------------
    # Identificação de fornecedor pelo email do remetente
    # ------------------------------------------------------------------

    def identificar_fornecedor(self, email_de: str):
        """
        Devolve o Fornecedor pelo remetente.
        Tenta por ordem: 1) email exacto; 2) domínio base (@kone → KONE).
        """
        if not self._fornecedor_repo:
            return None
        _, addr = parseaddr(email_de)
        addr = addr.strip().lower()
        if not addr:
            return None

        fornecedores = self._fornecedor_repo.list_all()

        # 1. Email exacto
        for f in fornecedores:
            emails_f = {(e or "").strip().lower()
                        for e in (f.email, f.comercial_email, f.administrativo_email)
                        if e}
            if addr in emails_f:
                return f

        dominio = _base_domain(addr)
        if not dominio:
            return None

        # 2. Mapa de associações definidas pelo utilizador
        if self._email_map_repo:
            fid = self._email_map_repo.find_by_domain(dominio)
            if fid:
                forn = self._fornecedor_repo.get_by_id(fid)
                if forn:
                    return forn

        # 3. Heurística de domínio base nos emails do fornecedor
        for f in fornecedores:
            for e in (f.email, f.comercial_email, f.administrativo_email):
                if e and _base_domain(e.strip().lower()) == dominio:
                    return f

        return None

    # ------------------------------------------------------------------
    # Catalogação — varrer pasta e produzir XLSX
    # ------------------------------------------------------------------

    def catalogar_pasta_faturas(self, pasta: str, limite: int = 500,
                                 callback=None) -> dict:
        """Delega no ImapClient. Devolve {"com_pdf", "com_link", "ignorados", "total"}."""
        try:
            return self._imap.catalogar_pasta_faturas(pasta, limite, callback)
        except Exception:
            logger.exception("Erro ao catalogar pasta '%s'", pasta)
            return {"com_pdf": [], "com_link": [], "ignorados": 0, "total": 0}

    def guardar_catalogo_xlsx(self, dados: dict, destino: Path) -> Path:
        """Guarda o catálogo de faturas num XLSX com timestamp. Devolve o caminho."""
        from datetime import datetime
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        def _forn(de: str) -> str:
            f = self.identificar_fornecedor(de)
            return f.nome if f else "—"

        def _autofit(ws):
            fill = PatternFill(fill_type="solid", fgColor="CADAD9")
            for cell in ws[1]:
                cell.font      = Font(bold=True)
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                w = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 70)

        destino.mkdir(parents=True, exist_ok=True)
        ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
        caminho = destino / f"catalogo_faturas_{ts}.xlsx"

        rows_pdf = [
            {
                "Fornecedor":   r.get("fornecedor") or _forn(r.get("de", "")),
                "De":           r.get("de", ""),
                "Assunto":      r.get("assunto", ""),
                "Data":         r.get("data", ""),
                "Ficheiros PDF": ", ".join(p["nome"] for p in r.get("pdfs", [])),
            }
            for r in dados.get("com_pdf", [])
        ]
        rows_link = [
            {
                "Fornecedor": r.get("fornecedor") or _forn(r.get("de", "")),
                "De":         r.get("de", ""),
                "Assunto":    r.get("assunto", ""),
                "Data":       r.get("data", ""),
                "URL(s)":     " | ".join(r.get("urls", [])),
            }
            for r in dados.get("com_link", [])
        ]

        cols_pdf  = ["Fornecedor", "De", "Assunto", "Data", "Ficheiros PDF"]
        cols_link = ["Fornecedor", "De", "Assunto", "Data", "URL(s)"]

        with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
            pd.DataFrame(rows_pdf  or [dict.fromkeys(cols_pdf,  "")]).to_excel(
                writer, index=False, sheet_name="Faturas PDF")
            pd.DataFrame(rows_link or [dict.fromkeys(cols_link, "")]).to_excel(
                writer, index=False, sheet_name="Links Faturas")
            _autofit(writer.sheets["Faturas PDF"])
            _autofit(writer.sheets["Links Faturas"])

        logger.info("Catálogo guardado: %s", caminho)
        return caminho

    def catalogar_e_guardar(self, pasta: str, destino: Path,
                            limite: int = 500, callback=None,
                            on_nao_reconhecido=None) -> dict:
        """
        Num único passo IMAP:
        1. Fetch completo da pasta (uma só ligação).
        2. Identifica o fornecedor por email exacto → domínio → NIF do PDF.
           Quando falha os três, chama on_nao_reconhecido(assunto, dominio)
           para que a GUI possa pedir orientação ao utilizador.
        3. Grava os PDFs em `destino` com prefixo do fornecedor.
        4. Grava XLSX de catálogo em `destino`.
        """
        from infrastructure.email.fatura_parser import extrair_texto_pdf, parse_fatura

        destino.mkdir(parents=True, exist_ok=True)

        # ── Fase A: fetch IMAP ────────────────────────────────────────────────
        dados = self._imap.catalogar_pasta_faturas(pasta, limite, callback)

        # ── Fase B: identificação (sem chamadas IMAP; GUI pode interagir) ─────
        for r in dados.get("com_pdf", []):
            forn      = None
            n_fatura  = None
            data_pdf  = None

            # 1. PDF: NIF → fornecedor
            for pdf in r.get("pdfs", []):
                try:
                    campos   = parse_fatura(extrair_texto_pdf(pdf["dados"]),
                                            nif_proprio=self._nif_proprio)
                    n_fatura = n_fatura or campos.get("n_fatura")
                    data_pdf = data_pdf or campos.get("data")
                    nif = campos.get("nif")
                    if nif and self._fornecedor_repo and not forn:
                        forn = self._fornecedor_repo.find_by_nif(nif)
                except Exception:
                    pass

            # 2. Histórico pela série do n.º fatura
            if not forn and n_fatura and self._faturacao_repo:
                nif_h, _ = self._faturacao_repo.buscar_emitente_por_serie(n_fatura)
                if nif_h and self._fornecedor_repo:
                    forn = self._fornecedor_repo.find_by_nif(nif_h)

            # 3. Domínio do email
            if not forn:
                forn = self.identificar_fornecedor(r.get("de", ""))

            # 4. Callback para o utilizador
            if not forn and on_nao_reconhecido:
                _, addr_raw = parseaddr(r.get("de", ""))
                addr_raw = addr_raw.strip().lower()
                dominio  = _base_domain(addr_raw)
                forn = on_nao_reconhecido(r.get("assunto", ""), dominio)
                if forn and dominio and self._email_map_repo:
                    self._email_map_repo.save(dominio, forn.id, addr_raw)

            r["_forn"]      = forn
            r["_n_fatura"]  = n_fatura
            r["_data_pdf"]  = data_pdf
            r["fornecedor"] = forn.nome if forn else "—"

        for r in dados.get("com_link", []):
            forn = self.identificar_fornecedor(r.get("de", ""))
            r["fornecedor"] = forn.nome if forn else "—"

        # ── Fase C: guardar PDFs ──────────────────────────────────────────────
        from infrastructure.email.fatura_parser import parse_date_iso as _pdi

        guardados        = 0
        nao_reconhecidos = 0

        for r in dados.get("com_pdf", []):
            forn     = r.pop("_forn", None)
            n_fatura = r.pop("_n_fatura", None)
            data_pdf = r.pop("_data_pdf", None)

            try:
                data_email_str = parsedate_to_datetime(r.get("data", "")).strftime("%Y%m%d")
            except Exception:
                data_email_str = "00000000"

            data_str = data_email_str
            if data_pdf:
                iso, _, _ = _pdi(data_pdf)
                if iso:
                    data_str = iso.replace("-", "")

            if not forn:
                nao_reconhecidos += 1

            pdfs = r.get("pdfs", [])
            for idx, pdf in enumerate(pdfs, 1):
                sufixo = f"_{idx}" if len(pdfs) > 1 else ""
                if forn:
                    tipo = _tipo_doc(n_fatura)
                    nome = f"{tipo}_{_prefixo_fornecedor(forn.nome)}_{data_str}{sufixo}.pdf"
                else:
                    nome = f"DESCON_{data_str}{sufixo}.pdf"
                caminho = destino / nome
                n = 1
                while caminho.exists():
                    stem = nome[:-4]
                    caminho = destino / f"{stem}_{n}.pdf"
                    n += 1
                try:
                    caminho.write_bytes(pdf["dados"])
                    guardados += 1
                    logger.info("PDF guardado: %s  (%d bytes)", caminho.name, len(pdf["dados"]))
                except Exception:
                    logger.exception("Erro ao guardar '%s'", nome)

        dados["guardados"]        = guardados
        dados["nao_reconhecidos"] = nao_reconhecidos

        # ── Fase D: XLSX ──────────────────────────────────────────────────────
        try:
            dados["xlsx"] = self.guardar_catalogo_xlsx(dados, destino)
        except Exception:
            logger.exception("Erro ao guardar catálogo XLSX")
            dados["xlsx"] = None

        return dados

    # ------------------------------------------------------------------
    # Fase 1 — Guardar PDFs em disco
    # ------------------------------------------------------------------

    def guardar_pdfs_pasta(self, pasta: str, destino: Path,
                           limite: int = 200, callback=None,
                           on_sem_dados=None) -> dict:
        """
        Descarrega PDFs de faturas da pasta IMAP e guarda-os em `destino`.
        Nomeia cada ficheiro: {fat|nc}_{5chars_fornecedor}_{YYYYMMDD}[_N].pdf

        Fluxo por PDF:
          1. Lê o PDF → extrai n_fatura, nif, data, total, iva
          2. Preenche campos em falta consultando o histórico de faturação
             (ex: NIF por série de n.º fatura já conhecida)
          3. Se ainda faltarem dados críticos → chama on_sem_dados(pdf_nome, campos)
             para o utilizador preencher via GUI
          4. Nomeia com base nos dados completos; usa DESCON se NIF nunca resolvido

        on_sem_dados(pdf_nome, campos) → dict | None
            Recebe nome do PDF e campos extraídos até ao momento.
            Devolve dict com campos corrigidos/preenchidos, ou None para saltar.

        Devolve {"guardados", "ignorados", "nao_reconhecidos", "ficheiros"}.
        """
        from infrastructure.email.fatura_parser import (
            extrair_texto_pdf, parse_fatura, parse_date_iso,
        )

        destino.mkdir(parents=True, exist_ok=True)
        emails_pdf = self._imap.obter_emails_com_pdf(pasta, limite)

        validos   = [em for em in emails_pdf
                     if not re.search(r"\b(?:recibo|receipt)\b",
                                      em["assunto"], re.IGNORECASE)]
        ignorados = len(emails_pdf) - len(validos)
        if ignorados:
            logger.info("Pasta «%s»: %d email(s) ignorado(s) (recibos).", pasta, ignorados)

        guardados        = 0
        ja_existem       = 0
        nao_reconhecidos = 0
        ficheiros        = []
        vistos_sessao: set[tuple[str, str]] = set()   # (n_fatura, nif) desta sessão

        for i, em in enumerate(validos, 1):
            if callback:
                callback(i, len(validos), em["assunto"])

            try:
                data_email_str = parsedate_to_datetime(em["data"]).strftime("%Y%m%d")
            except Exception:
                data_email_str = "00000000"

            uid  = em["uid"]
            pdfs = em["pdfs"]

            for idx, pdf in enumerate(pdfs, 1):
                pdf_nome = pdf.get("nome", f"anexo_{uid}_{idx}.pdf")

                # ── 1. Leitura do PDF ─────────────────────────────────────────
                campos: dict = {}
                try:
                    campos = parse_fatura(
                        extrair_texto_pdf(pdf["dados"]),
                        nif_proprio=self._nif_proprio,
                    )
                except Exception:
                    logger.exception("Erro ao parsear '%s'", pdf_nome)

                # ── 2. Histórico: preencher NIF em falta pela série do n.º ────
                if not campos.get("nif") and campos.get("n_fatura") and self._faturacao_repo:
                    nif_h, nome_h = self._faturacao_repo.buscar_emitente_por_serie(
                        campos["n_fatura"]
                    )
                    if nif_h:
                        campos["nif"] = nif_h
                        logger.info("NIF preenchido via histórico (%s): %s",
                                    campos["n_fatura"][:10], nif_h)

                # ── 3. Pedir ao utilizador se NIF ainda em falta ─────────────
                if not campos.get("nif") and on_sem_dados:
                    correcao = on_sem_dados(pdf_nome, dict(campos))
                    if correcao:
                        campos.update(correcao)

                # ── Resolver fornecedor pelo NIF ─────────────────────────────
                nif  = campos.get("nif")
                forn = None
                if nif and self._fornecedor_repo:
                    forn = self._fornecedor_repo.find_by_nif(nif)

                # ── Deduplicação: sessão corrente + histórico em BD ───────────
                n_fatura_campo = campos.get("n_fatura")
                if n_fatura_campo and nif:
                    chave = (n_fatura_campo, nif)
                    if chave in vistos_sessao:
                        logger.info("Duplicado na sessão, ignorado: %s / %s",
                                    n_fatura_campo, nif)
                        ja_existem += 1
                        continue
                    if (self._faturacao_repo
                            and self._faturacao_repo.existe(n_fatura_campo, nif)):
                        logger.info("Fatura já existe na BD, ignorada: %s / %s",
                                    n_fatura_campo, nif)
                        ja_existem += 1
                        continue
                    vistos_sessao.add(chave)

                # ── Data para o nome: da fatura ou do email ───────────────────
                data_str = data_email_str
                if campos.get("data"):
                    iso, _, _ = parse_date_iso(campos["data"])
                    if iso:
                        data_str = iso.replace("-", "")

                # ── Construir nome do ficheiro ────────────────────────────────
                sufixo = f"_{idx}" if len(pdfs) > 1 else ""
                if forn:
                    tipo   = _tipo_doc(n_fatura_campo)
                    nome   = f"{tipo}_{_prefixo_fornecedor(forn.nome)}_{data_str}{sufixo}.pdf"
                    logger.info("Fornecedor: %s → %s", forn.nome, nome)
                else:
                    nome = f"DESCON_{data_str}{sufixo}.pdf"
                    nao_reconhecidos += 1
                    logger.info("Fornecedor não identificado para '%s' (%s)", pdf_nome, em["de"])

                caminho = destino / nome
                n = 1
                while caminho.exists():
                    stem = nome[:-4]   # sem .pdf
                    caminho = destino / f"{stem}_{n}.pdf"
                    n += 1
                try:
                    caminho.write_bytes(pdf["dados"])
                    guardados += 1
                    ficheiros.append(caminho)
                    logger.info("PDF guardado: %s  (%d bytes)", caminho.name, len(pdf["dados"]))
                except Exception:
                    logger.exception("Erro ao guardar '%s'", nome)

        return {"guardados": guardados, "ignorados": ignorados,
                "ja_existem": ja_existem,
                "nao_reconhecidos": nao_reconhecidos, "ficheiros": ficheiros}

    # ------------------------------------------------------------------
    # Fase 2 — Ler PDFs do disco e extrair campos de fatura
    # ------------------------------------------------------------------

    def ler_faturas_directorio(self, directorio: Path, callback=None,
                               on_dados_incompletos=None) -> list[dict]:
        """Extrai campos de fatura de todos os PDFs no directório.

        on_dados_incompletos(nome_pdf, fornecedor, campos) → dict | None
            Chamada quando qualquer campo visível estiver em falta,
            no máximo uma vez por fornecedor (por NIF) em cada sessão.
            Devolve dict com correções ou None para manter o que foi extraído.
        """
        from infrastructure.email.fatura_parser import extrair_texto_pdf, parse_fatura, log_fatura

        pdfs = sorted(directorio.glob("*.pdf"))
        resultados = []
        # serie[:10] → NIF fornecido pelo user; evita perguntar várias vezes pelo mesmo fornecedor
        _nif_por_serie: dict[str, str] = {}
        # NIFs já tratados (via diálogo ou cache) — não volta a perguntar
        _nifs_tratados: set[str] = set()

        for i, caminho in enumerate(pdfs, 1):
            if callback:
                callback(i, len(pdfs), caminho.name)
            try:
                texto  = caminho.read_bytes()
                if not texto:
                    raise ValueError("ficheiro vazio")
                campos = parse_fatura(extrair_texto_pdf(texto), nif_proprio=self._nif_proprio)
            except Exception:
                logger.exception("Erro ao processar '%s'", caminho.name)
                campos = {"n_fatura": None, "data": None, "vencimento": None,
                          "nif": None, "total": None, "iva": None}

            # Injectar NIF da cache de série se ainda não foi encontrado pelo parser
            if not campos.get("nif") and campos.get("n_fatura"):
                serie = campos["n_fatura"][:10]
                if serie in _nif_por_serie:
                    campos["nif"] = _nif_por_serie[serie]

            _incompleto = (
                not campos.get("n_fatura") or
                not campos.get("nif")      or
                campos.get("total")     is None or
                not campos.get("data")     or
                not campos.get("vencimento") or
                campos.get("iva") is None
            )
            nif_actual = campos.get("nif")
            # Só pergunta se ainda incompleto E este fornecedor ainda não foi tratado
            if _incompleto and on_dados_incompletos and nif_actual not in _nifs_tratados:
                try:
                    forn = None
                    if nif_actual and self._fornecedor_repo:
                        forn = self._fornecedor_repo.find_by_nif(nif_actual)
                    correcao = on_dados_incompletos(
                        caminho.name,
                        forn.nome if forn else "—",
                        {**campos},
                    )
                    if correcao:
                        campos.update(correcao)
                        nif_actual = campos.get("nif")
                        # Guardar NIF na cache de série e marcar como tratado
                        if nif_actual and campos.get("n_fatura"):
                            _nif_por_serie[campos["n_fatura"][:10]] = nif_actual
                        if nif_actual:
                            _nifs_tratados.add(nif_actual)
                except Exception:
                    logger.exception("Erro no diálogo para '%s'", caminho.name)
            elif nif_actual:
                _nifs_tratados.add(nif_actual)

            log_fatura(caminho.name, str(directorio), caminho.stem, "", campos)
            resultados.append({
                "pdf_nome": caminho.name,
                "assunto":  caminho.stem,
                **campos,
            })
        return resultados

    def ler_faturas_pasta(self, pasta: str, limite: int = 200, callback=None) -> dict:
        """
        Lê faturas em PDF da pasta indicada.
        Filtra emails cujo assunto contenha 'recibo' ou 'receipt'.
        Devolve {"resultados": [...], "ignorados": int}.
        """
        from infrastructure.email.fatura_parser import extrair_texto_pdf, parse_fatura, log_fatura

        emails_pdf = self._imap.obter_emails_com_pdf(pasta, limite)

        resultados = []
        ignorados  = 0
        validos    = [
            em for em in emails_pdf
            if not re.search(r"\b(?:recibo|receipt)\b", em["assunto"], re.IGNORECASE)
        ]
        ignorados = len(emails_pdf) - len(validos)
        if ignorados:
            logger.info("Pasta «%s»: %d email(s) ignorado(s) por serem recibos.", pasta, ignorados)

        total = len(validos)
        for i, em in enumerate(validos, 1):
            if callback:
                callback(i, total, em["assunto"])
            for pdf in em["pdfs"]:
                try:
                    texto  = extrair_texto_pdf(pdf["dados"])
                    campos = parse_fatura(texto, nif_proprio=self._nif_proprio)
                    if not texto.strip():
                        logger.warning("PDF '%s': sem texto extraível (possível imagem digitalizada).", pdf["nome"])
                except Exception:
                    logger.exception("Erro ao processar PDF '%s'", pdf["nome"])
                    campos = {"n_fatura": None, "data": None, "nif": None, "total": None, "iva": None}

                log_fatura(pdf["nome"], pasta, em["assunto"], em["de"], campos)

                resultados.append({
                    "pasta":      pasta,
                    "uid":        em["uid"],
                    "de":         em["de"],
                    "assunto":    em["assunto"],
                    "data_email": em["data"],
                    "pdf_nome":   pdf["nome"],
                    **campos,
                })

        return {"resultados": resultados, "ignorados": ignorados}

    # ------------------------------------------------------------------
    # Fase 3 — Recolher dados: PDFs do disco → faturacao_documentos
    # ------------------------------------------------------------------

    def recolher_faturas_directorio(self, directorio: Path,
                                    faturacao_repo, callback=None,
                                    on_dados_incompletos=None) -> dict:
        """
        Lê PDFs de `directorio`, extrai campos de fatura e persiste em
        faturacao_documentos.

        on_dados_incompletos(nome_pdf, fornecedor, campos) → dict | None
            Chamada quando n_fatura, nif ou total não são detectados.
            Recebe o nome do PDF, o nome do fornecedor (ou "—") e os campos
            até então extraídos. Deve devolver um dict com os campos corrigidos,
            ou None para ignorar a fatura.

        Devolve {"resultados": [...], "importados": int, "ja_existem": int,
                 "sem_dados": int}.
        """
        from infrastructure.email.fatura_parser import (
            extrair_texto_pdf, parse_fatura, parse_date_iso,
        )

        pdfs = sorted(directorio.glob("*.pdf"))
        resultados = []
        importados = 0
        ja_existem = 0
        sem_dados  = 0
        # serie[:10] → NIF fornecido pelo user nesta sessão
        _nif_por_serie: dict[str, str] = {}
        # NIFs já tratados — não volta a perguntar pelo mesmo fornecedor
        _nifs_tratados: set[str] = set()

        for i, caminho in enumerate(pdfs, 1):
            if callback:
                callback(i, len(pdfs), caminho.name)

            try:
                campos = parse_fatura(extrair_texto_pdf(caminho.read_bytes()), nif_proprio=self._nif_proprio)
            except Exception:
                logger.exception("Erro ao processar '%s'", caminho.name)
                campos = {}

            n_fatura   = campos.get("n_fatura")
            nif        = campos.get("nif")

            # Injectar NIF da cache de série se ainda não foi encontrado pelo parser
            if not nif and n_fatura:
                serie = n_fatura[:10]
                if serie in _nif_por_serie:
                    nif = _nif_por_serie[serie]
                    campos["nif"] = nif

            total      = campos.get("total")
            iva        = campos.get("iva") or 0
            data_str   = campos.get("data")
            vencimento = campos.get("vencimento")

            # Identificar fornecedor pelo NIF (mesmo antes de pedir ajuda)
            forn = None
            if nif and self._fornecedor_repo:
                forn = self._fornecedor_repo.find_by_nif(nif)

            # Disparar diálogo apenas se incompleto E fornecedor ainda não tratado nesta sessão
            _incompleto = (
                not campos.get("n_fatura") or
                not campos.get("nif")      or
                campos.get("total")     is None or
                not campos.get("data")     or
                not campos.get("vencimento") or
                campos.get("iva") is None
            )
            if _incompleto and on_dados_incompletos and nif not in _nifs_tratados:
                try:
                    correcao = on_dados_incompletos(
                        caminho.name,
                        forn.nome if forn else "—",
                        {**campos},
                    )
                except Exception:
                    logger.exception("Erro no diálogo de dados em falta para '%s'", caminho.name)
                    correcao = None
                if correcao:
                    campos.update(correcao)
                    n_fatura   = campos.get("n_fatura")
                    nif        = campos.get("nif")
                    total      = campos.get("total")
                    iva        = campos.get("iva") or 0
                    data_str   = campos.get("data")
                    vencimento = campos.get("vencimento")
                    if nif and self._fornecedor_repo and not forn:
                        forn = self._fornecedor_repo.find_by_nif(nif)
                    # Guardar NIF na cache de série e marcar como tratado
                    if nif and n_fatura:
                        _nif_por_serie[n_fatura[:10]] = nif
                    if nif:
                        _nifs_tratados.add(nif)
            elif nif:
                _nifs_tratados.add(nif)

            if not n_fatura or not nif or total is None:
                sem_dados += 1
                resultados.append({
                    "pdf_nome": caminho.name, **campos,
                    "estado": "Sem dados", "fornecedor": forn.nome if forn else "—",
                })
                continue

            # Tipo de documento a partir do prefixo do n.º
            tipo_m = re.match(
                r"^(FT|FR|FS|FA|NC|ND|RC|RG|VD|ORC|GT|GR)", n_fatura, re.IGNORECASE
            )
            tipo = tipo_m.group(1).upper() if tipo_m else None

            data_iso, ano, mes     = parse_date_iso(data_str)
            venc_iso, _, _         = parse_date_iso(vencimento)
            base = round(total - iva, 2)

            record = {
                "fornecedor_id":    forn.id if forn else None,
                "nif_emitente":     nif,
                "nome_emitente":    forn.nome if forn else None,
                "numero_documento": n_fatura,
                "tipo":             tipo,
                "data_emissao":     data_iso,
                "ano":              ano,
                "mes":              mes,
                "base_tributavel":  base,
                "iva":              iva,
                "total":            total,
                "data_vencimento":  venc_iso,
                "situacao":         "Registado",
            }

            if faturacao_repo.existe(n_fatura, nif):
                ja_existem += 1
                estado = "Já existe"
            else:
                faturacao_repo.inserir(record)
                importados += 1
                estado = "Importado"

            resultados.append({
                "pdf_nome":       caminho.name,
                **campos,
                "vencimento":     vencimento,
                "fornecedor":     forn.nome if forn else "—",
                "estado":         estado,
            })

        return {
            "resultados": resultados,
            "importados": importados,
            "ja_existem": ja_existem,
            "sem_dados":  sem_dados,
        }

    def listar_mensagens(self, pasta: str, limite: int = 200) -> list[dict]:
        self._pasta_actual = pasta
        try:
            return self._imap.listar_mensagens(pasta, limite)
        except Exception:
            logger.exception("Erro ao listar mensagens de '%s'", pasta)
            return []

    def obter_mensagem(self, uid: str) -> dict | None:
        if not self._pasta_actual:
            return None
        try:
            return self._imap.obter_mensagem(self._pasta_actual, uid)
        except Exception:
            logger.exception("Erro ao obter mensagem uid=%s", uid)
            return None
