# presentation/correio/faturas_gui.py
#
# Fase 2 — Ler PDFs de faturas guardados em disco, extrair campos e exportar.
#
# Lê todos os *.pdf de controller.pasta_faturas, aplica o parser e
# apresenta os resultados numa tabela. Exporta para XLSX via pandas.

import json
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import ttk, filedialog

from presentation.shared.base_gui import BaseGui as BG

_CORRECOES_JSON = Path(__file__).parent.parent.parent / "tools" / "correcoes_faturas.json"

_COLS = [
    ("pdf_nome",   "Ficheiro PDF",   140, "w"),
    ("n_fatura",   "N.º Fatura",     110, "w"),
    ("data",       "Data Emissão",    88, "center"),
    ("vencimento", "Data Pagamento",  88, "center"),
    ("nif",        "NIF",             95, "center"),
    ("total",      "Total (€)",       82, "e"),
    ("iva",        "IVA (€)",         72, "e"),
]

_CAMPOS_DLG = [
    ("n_fatura",   "N.º Fatura"),
    ("nif",        "NIF emitente"),
    ("data",       "Data emissão"),
    ("vencimento", "Data pagamento"),
    ("total",      "Total (€)"),
    ("iva",        "IVA (€)"),
]


# ------------------------------------------------------------------
# Persistência de correções (partilhada com debug_faturas.py e recolher_dados_gui.py)
# ------------------------------------------------------------------

def _guardar_correcao(pdf_nome: str, campo: str, extraido, correto: str) -> None:
    dados: list = []
    if _CORRECOES_JSON.exists():
        try:
            dados = json.loads(_CORRECOES_JSON.read_text(encoding="utf-8"))
        except Exception:
            pass
    dados.append({
        "pdf":      pdf_nome,
        "campo":    campo,
        "extraido": str(extraido) if extraido is not None else None,
        "correto":  correto,
        "data":     datetime.now().isoformat(timespec="seconds"),
    })
    _CORRECOES_JSON.parent.mkdir(parents=True, exist_ok=True)
    _CORRECOES_JSON.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


class FaturasGUI(BG):

    def __init__(self, parent, controller):
        self._resultados = []
        super().__init__(parent, controller)

    def _post_init(self):
        self.build_menu_buttons([
            ("Analisar PDFs", self._analisar),
            ("Exportar XLSX", self._exportar),
        ])
        self._build()
        self._mostrar_directorio()

    # ------------------------------------------------------------------
    # Layout
    # ------------------------------------------------------------------

    def _build(self):
        outer = tk.Frame(self.frame, bg=self.BG)
        outer.pack(fill="both", expand=True, padx=8, pady=8)
        outer.rowconfigure(1, weight=1)
        outer.columnconfigure(0, weight=1)

        destino = str(self.controller.pasta_faturas) if self.controller.pasta_faturas else "—"
        tk.Label(
            outer, text=f"Pasta:  {destino}",
            bg=self.BG, fg=self.FG, font=("Segoe UI", 9, "bold"), anchor="w",
        ).grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 6))

        self.tree = ttk.Treeview(
            outer, columns=[c[0] for c in _COLS],
            show="headings", selectmode="browse",
        )
        for key, label, width, anchor in _COLS:
            self.tree.heading(key, text=label)
            self.tree.column(key, width=width, anchor=anchor, minwidth=40,
                             stretch=(key in ("pdf_nome", "n_fatura")))

        self.tree.tag_configure("ok",      foreground="#1a7a1a")
        self.tree.tag_configure("parcial", foreground="#8a5c00")
        self.tree.tag_configure("vazio",   foreground="#999999")

        sb_v = ttk.Scrollbar(outer, orient="vertical",   command=self.tree.yview)
        sb_h = ttk.Scrollbar(outer, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)
        self.tree.grid(row=1, column=0, sticky="nsew")
        sb_v.grid(row=1, column=1, sticky="ns")
        sb_h.grid(row=2, column=0, sticky="ew")

        self.lbl_status = tk.Label(
            outer, text="Prima «Analisar PDFs» para iniciar.",
            bg=self.BG, fg=self.FG, font=("Segoe UI", 8), anchor="w",
        )
        self.lbl_status.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(4, 0))

    # ------------------------------------------------------------------

    def _mostrar_directorio(self):
        d = self.controller.pasta_faturas
        if d and d.exists():
            n = len(list(d.glob("*.pdf")))
            self._set_status(f"{n} PDF(s) encontrado(s) em {d}. Prima «Analisar PDFs».")
        elif d:
            self._set_status(f"Pasta ainda não existe: {d}")
        else:
            self._set_status("pasta_faturas não configurada.")

    # ------------------------------------------------------------------
    # Análise
    # ------------------------------------------------------------------

    def _analisar(self):
        directorio = self.controller.pasta_faturas
        if not directorio or not directorio.exists():
            self.informuser("Pasta não encontrada",
                            f"A pasta não existe:\n{directorio}", "warning")
            return

        pdfs = list(directorio.glob("*.pdf"))
        if not pdfs:
            self.informuser("Sem PDFs", "Não foram encontrados ficheiros PDF.", "warning")
            return

        self.tree.delete(*self.tree.get_children())
        self._resultados = []

        def progresso(i, total, nome):
            self._set_status(f"{i}/{total}: {nome}")
            self.frame.update_idletasks()

        self._resultados = self.controller.ler_faturas_directorio(
            directorio,
            callback=progresso,
            on_dados_incompletos=self._pedir_dados_em_falta,
        )

        for r in self._resultados:
            campos_ok = sum(
                1 for k in ("n_fatura", "data", "vencimento", "nif", "total", "iva")
                if r.get(k) is not None
            )
            tag = "ok" if campos_ok >= 5 else ("parcial" if campos_ok >= 3 else "vazio")
            self.tree.insert("", "end", tags=(tag,), values=(
                r["pdf_nome"],
                r.get("n_fatura")   or "—",
                r.get("data")       or "—",
                r.get("vencimento") or "—",
                r.get("nif")        or "—",
                f"{r['total']:.2f}" if r.get("total") is not None else "—",
                f"{r['iva']:.2f}"   if r.get("iva")   is not None else "—",
            ))

        total     = len(self._resultados)
        com_nif   = sum(1 for r in self._resultados if r.get("nif"))
        com_total = sum(1 for r in self._resultados if r.get("total") is not None)
        self._set_status(
            f"{total} PDF(s)  |  {com_nif} com NIF  |  {com_total} com total identificado"
        )

    # ------------------------------------------------------------------
    # Diálogo de preenchimento manual
    # ------------------------------------------------------------------

    def _pedir_dados_em_falta(self, nome_pdf: str, fornecedor: str, campos: dict) -> dict | None:
        """
        Modal que pede ao utilizador os campos não detectados automaticamente.
        Campos em falta aparecem a vermelho.
        Correções são gravadas em tools/correcoes_faturas.json para aprendizagem.
        """
        root = self.frame.winfo_toplevel()
        dlg  = tk.Toplevel(root)
        dlg.title("Dados em falta — ajuda necessária")
        dlg.resizable(False, False)
        dlg.transient(root)

        pad = 14

        tk.Label(
            dlg,
            text="Não foi possível detectar automaticamente todos os campos.",
            font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w", padx=pad, pady=(12, 2))

        info = tk.Frame(dlg)
        info.pack(fill="x", padx=pad, pady=(0, 6))
        for row, (label, val) in enumerate([
            ("PDF:",         nome_pdf[:80]),
            ("Fornecedor:",  fornecedor[:80]),
        ]):
            tk.Label(info, text=label, font=("Segoe UI", 9, "bold"),
                     width=12, anchor="w").grid(row=row, column=0, sticky="w")
            tk.Label(info, text=val, font=("Segoe UI", 9),
                     wraplength=380, justify="left").grid(row=row, column=1, sticky="w")

        tk.Frame(dlg, height=1, bg="#cccccc").pack(fill="x", padx=pad, pady=(0, 8))

        entries: dict[str, tk.StringVar] = {}
        form = tk.Frame(dlg)
        form.pack(fill="x", padx=pad)
        primeiro_vazio = None

        for row, (chave, label) in enumerate(_CAMPOS_DLG):
            val_atual = campos.get(chave)
            em_falta  = val_atual is None or (isinstance(val_atual, str) and not val_atual.strip())
            cor    = "#cc0000" if em_falta else "#333333"
            sufixo = " *" if em_falta else ""

            tk.Label(
                form,
                text=label + sufixo,
                font=("Segoe UI", 9, "bold" if em_falta else "normal"),
                fg=cor, width=16, anchor="w",
            ).grid(row=row, column=0, sticky="w", pady=2)

            var = tk.StringVar(value=str(val_atual) if val_atual is not None else "")
            ent = tk.Entry(form, textvariable=var, width=34, font=("Segoe UI", 9))
            ent.grid(row=row, column=1, sticky="ew", pady=2)
            entries[chave] = var

            if em_falta and primeiro_vazio is None:
                primeiro_vazio = ent

        tk.Label(
            dlg,
            text="* campo em falta  |  Enter = confirmar  |  Esc = ignorar",
            font=("Segoe UI", 8), fg="#888888",
        ).pack(anchor="w", padx=pad, pady=(6, 0))

        resultado = [None]

        def _confirmar():
            correcao = {}
            for chave, var in entries.items():
                v = var.get().strip()
                if not v:
                    continue
                val_orig = campos.get(chave)
                if str(val_orig or "") != v:
                    _guardar_correcao(nome_pdf, chave, val_orig, v)
                if chave in ("total", "iva"):
                    try:
                        correcao[chave] = float(v.replace(",", "."))
                    except ValueError:
                        pass
                else:
                    correcao[chave] = v
            resultado[0] = correcao if correcao else {}
            dlg.destroy()

        def _ignorar():
            dlg.destroy()

        btns = tk.Frame(dlg)
        btns.pack(fill="x", padx=pad, pady=(10, 14))
        tk.Button(btns, text="Usar estes dados", command=_confirmar,
                  font=("Segoe UI", 9), width=16).pack(side="left", padx=(0, 8))
        tk.Button(btns, text="Ignorar",          command=_ignorar,
                  font=("Segoe UI", 9), width=10).pack(side="left")

        dlg.bind("<Return>", lambda _: _confirmar())
        dlg.bind("<Escape>", lambda _: _ignorar())

        # Centrar, garantir visibilidade, depois grab (ordem importa no Windows)
        dlg.update_idletasks()
        rx = root.winfo_rootx() + (root.winfo_width()  - dlg.winfo_width())  // 2
        ry = root.winfo_rooty() + (root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{rx}+{ry}")
        dlg.deiconify()
        dlg.attributes("-topmost", True)
        dlg.update()
        dlg.grab_set()
        dlg.attributes("-topmost", False)
        if primeiro_vazio:
            primeiro_vazio.focus_set()

        dlg.wait_window()
        return resultado[0]

    # ------------------------------------------------------------------
    # Exportação
    # ------------------------------------------------------------------

    def _exportar(self):
        if not self._resultados:
            self.informuser("Sem dados", "Analise os PDFs antes de exportar.", "warning")
            return

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar faturas",
            initialfile="faturas.xlsx",
        )
        if not caminho:
            return
        try:
            _exportar_xlsx(self._resultados, Path(caminho))
            self.informuser("Exportado", f"Ficheiro guardado:\n{caminho}")
        except Exception as exc:
            self.informuser("Erro ao exportar", str(exc), "error")

    # ------------------------------------------------------------------

    def _set_status(self, msg: str):
        self.lbl_status.config(text=msg)


# ------------------------------------------------------------------
# Exportação XLSX
# ------------------------------------------------------------------

def _exportar_xlsx(resultados: list[dict], caminho: Path) -> None:
    import pandas as pd

    colunas = {
        "pdf_nome":   "Ficheiro PDF",
        "n_fatura":   "N.º Fatura",
        "data":       "Data Emissão",
        "vencimento": "Data Pagamento",
        "nif":        "NIF Emissor",
        "total":      "Total (€)",
        "iva":        "IVA (€)",
    }
    df = pd.DataFrame(resultados)
    for k in colunas:
        if k not in df.columns:
            df[k] = None
    df = df[list(colunas)].rename(columns=colunas)

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Faturas")
        ws = writer.sheets["Faturas"]

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        fill = PatternFill(fill_type="solid", fgColor="CADAD9")
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 55)
        for row in ws.iter_rows(min_row=2):
            for cell in (row[5], row[6]):   # Total, IVA
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
