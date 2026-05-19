# infrastructure/email/analise_xlsx.py
#
# Exportação de análise da caixa de correio para XLSX.
# Sheets: Resumo | Emissores | Destinatários | Palavras Global

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HDR_FILL = PatternFill(fill_type="solid", fgColor="CADAD9")
_HDR_FONT = Font(bold=True)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cabecalho(ws, colunas: list[str]):
    ws.append(colunas)
    for cell in ws[1]:
        cell.font  = _HDR_FONT
        cell.fill  = _HDR_FILL
        cell.alignment = _HDR_ALIGN
    ws.row_dimensions[1].height = 28


def _auto_largura(ws, max_col: int = 80):
    for col in ws.columns:
        largura = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(largura + 4, max_col)


def exportar_analise_xlsx(dados: dict, caminho: Path) -> None:
    """
    dados  : resultado de CorreioController.analisar_todas_pastas()
    caminho: destino do ficheiro .xlsx
    """
    wb = Workbook()

    # ── Resumo por pasta ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    _cabecalho(ws, ["Pasta", "Total", "Com Anexos", "% Anexos", "Top 5 Palavras-chave"])
    for info in dados["por_pasta"]:
        total = info["total"]
        pct   = round(100 * info["num_anexos"] / total, 1) if total else 0
        palavras_str = ", ".join(w for w, _ in info["palavras_chave"])
        ws.append([info["pasta"], total, info["num_anexos"], pct, palavras_str])
    _auto_largura(ws)

    # ── Emissores por pasta ───────────────────────────────────────────
    ws_em = wb.create_sheet("Emissores")
    _cabecalho(ws_em, ["Pasta", "Emissor", "N"])
    for info in dados["por_pasta"]:
        for addr, n in info["emissores"]:
            ws_em.append([info["pasta"], addr, n])
    _auto_largura(ws_em)

    # ── Destinatários por pasta ───────────────────────────────────────
    ws_dest = wb.create_sheet("Destinatários")
    _cabecalho(ws_dest, ["Pasta", "Destinatário", "N"])
    for info in dados["por_pasta"]:
        for addr, n in info["destinatarios"]:
            ws_dest.append([info["pasta"], addr, n])
    _auto_largura(ws_dest)

    # ── Palavras global ───────────────────────────────────────────────
    ws_pal = wb.create_sheet("Palavras Global")
    _cabecalho(ws_pal, ["Palavra", "Ocorrências"])
    for w, n in dados["palavras_global"]:
        ws_pal.append([w, n])
    _auto_largura(ws_pal)

    wb.save(caminho)
