#atrpt/infrastructure/file_system/io.py

from pathlib import Path
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlwings as xw
from datetime import datetime
import shutil
import xlsxwriter
from typing import Union

def ler_excel(path: Path, **kwargs):
    if not isinstance(path, Path):
        raise TypeError("ler_excel espera Path absoluto")
    if not path.exists():
        raise FileNotFoundError(path)
    return pd.read_excel(path, **kwargs)

def ler_parquet(path: Path) -> pd.DataFrame:
    if not isinstance(path, Path):
        raise TypeError("ler_parquet espera Path absoluto")
    if not path.exists():
        raise FileNotFoundError(path)
    return pd.read_parquet(path)

def backup_file(src: Path, *, backup_dir: Path) -> Path:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    dst = backup_dir / f"{src.stem}_{ts}{src.suffix}"
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst

def guardar_excel_com_backup(df, path: Path, *, backup_dir: Path):
    backup_file(path, backup_dir=backup_dir)
    df.to_excel(path, index=False)

def ler_jpg(path: Path) -> Image.Image:
    if not isinstance(path, Path):
        raise TypeError("ler_jpg espera Path absoluto")    
    # Aceita str ou Path
    if not path.exists():
        raise TypeError("ler_jpg espera Path absoluto")
    img = Image.open(path)
    img.load()
    return img

def ler_csv(path: Path, **kwargs):
    if not isinstance(path, Path):
        raise TypeError("ler_csv espera Path absoluto")    
    if not path.exists():
        raise FileNotFoundError(p)
    return pd.read_csv(path, **kwargs)

def atualizar_excel_com_ligacoes(filepath):
    """
    Abre um ficheiro Excel e tenta atualizar ligações externas de forma segura.
    Ignora falhas de UpdateLink e fecha o Excel limpo.
    """
    import xlwings as xw
    if not os.path.exists(filepath):
        print(f"❌ Ficheiro não encontrado: {filepath}")
        return

    app = xw.App(visible=False)
    app.display_alerts = False
    app.screen_updating = False

    try:
        wb = app.books.open(filepath, update_links=False, read_only=False)
        try:
            links = wb.api.LinkSources()
            if links:
                for link in links:
                    try:
                        wb.api.UpdateLink(link)
                        print(f"🔗 Ligação atualizada: {link}")
                    except Exception as e:
                        print(f"⚠ Falha ao atualizar ligação {link}: {e}")
            else:
                print("ℹ Nenhuma ligação encontrada.")
        except Exception:
            print("⚠ Falha ao listar ligações (pode não haver nenhuma).")

        wb.save()
    except Exception as e:
        print(f"❌ Erro ao abrir {filepath}: {e}")
    finally:
        try:
            wb.close()
        except Exception:
            pass
        app.quit()

def ajusta_excel(file_name: str) -> None:
    """
    Ajusta aproximadamente a largura das colunas de todas as folhas
    de um ficheiro Excel (.xlsx / .xlsm) sem abrir o Excel.

    Baseia-se no tamanho máximo do conteúdo de cada coluna.
    """
    if not file_name:
        print("⚠ ajusta_excel: file_name vazio ou None.")
        return

    if not os.path.exists(file_name):
        print(f"⚠ ajusta_excel: ficheiro não encontrado -> {file_name}")
        return
    ext = os.path.splitext(file_name)[1].lower()
    extensoes_suportadas = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    if ext not in extensoes_suportadas:
        print(f"ℹ ajusta_excel: formato não suportado ({ext}), ignorado.")
        return
    try:
        wb = load_workbook(file_name)
        for ws in wb.worksheets:
            for col_cells in ws.columns:
                max_length = 0
                col = col_cells[0].column  # número da coluna
                col_letter = get_column_letter(col)

                for cell in col_cells:
                    value = cell.value
                    if value is not None:
                        length = len(str(value))
                        if length > max_length:
                            max_length = length

                if max_length > 0:
                    # pequeno padding para não cortar texto
                    ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_name)
        print(f"✔ ajusta_excel: colunas ajustadas em {file_name}")
    except Exception as e:
        print(f"⚠ ajusta_excel: erro em '{file_name}': {e}")

def escrever_excel(
    df: pd.DataFrame,
    path: Path,
    *,
    index: bool = False,
    engine: str | None = None,
    overwrite: bool = False,
) -> None:
    """
    Escreve um dataFrame para um ficheiro Excel e ajusta
    automaticamente a largura das colunas ao conteúdo.
    """
    if not isinstance(df, pd.DataFrame):
        raise TypeError(f"df deve ser pandas.dataFrame, recebido {type(df)}")
    if not isinstance(path, Path):
        raise TypeError(f"path deve ser pathlib.Path, recebido {type(path)}")

    path.parent.mkdir(parents=True, exist_ok=True)

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Extensão inválida para Excel: {path.suffix}")

    if path.exists() and not overwrite:
        raise FileExistsError(path)

    engine = engine or "xlsxwriter"

    with pd.ExcelWriter(path, engine=engine) as writer:
        df.to_excel(writer, index=index, sheet_name="Sheet1")

        worksheet = writer.sheets["Sheet1"]

        # ajuste automático das colunas
        for i, col in enumerate(df.columns):
            # largura = max(len(header), len(maior valor da coluna))
            max_len = len(str(col))

            if not df[col].empty:
                max_val_len = df[col].astype(str).map(len).max()
                if pd.notna(max_val_len):
                    max_len = max(max_len, int(max_val_len))

            # margem de segurança
            worksheet.set_column(i, i, max_len + 2)

def escrever_parquet(df: pd.DataFrame,path: Path,*,index: bool = False,overwrite: bool = False) -> None:
    """
    Escreve um dataFrame para um ficheiro parquet
    """
    if not isinstance(df, pd.DataFrame):
        raise TypeError(f"df deve ser pandas.dataFrame, recebido {type(df)}")
    if not isinstance(path, Path):
        raise TypeError(f"path deve ser pathlib.Path, recebido {type(path)}")
    # garantir diretório pai
    path.parent.mkdir(parents=True, exist_ok=True)
    # valida extensão
    if path.suffix.lower() not in {".parquet",}:
        raise ValueError(f"Extensão inválida para parquet: {path.suffix}")
    if path.exists() and not overwrite:
        raise FileExistsError(path)
    df.to_parquet(path, index=index)

def guardar_df( df, path):
    if df is None or df.empty:
        return
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        df.to_excel(path, index=False)
    elif path.suffix.lower() == ".parquet":
        df.to_parquet(path, index=False)
    else:
            raise ValueError(f"Formato não suportado: {path}")